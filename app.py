import pandas as pd
import os
from openpyxl.styles import colors, Font, PatternFill, Alignment, Border, Side
from openpyxl import load_workbook
from spire.xls import *
from spire.xls.common import *


"""

This piee code is used to generate a report for the customer's outstanding amount. 
The code reads the data from the excel files and then groups the data by the customer's name. 
The code then calculates the total amount of the invoice for each customer and then subtracts the total amount of the returns and payments made by the customer. 
The code then adds the outstanding amount of the customer to the final report. The code then writes the final report to the excel file.

"""

# import environment variables

global invoice_df
global payments_df
global returns_df


invoice_df = pd.read_excel("INVOICE.xlsx")
payments_df = pd.read_excel("PAYMENTS.xlsx")
returns_df = pd.read_excel("RETURNS.xlsx")
outstanding_df = pd.read_excel("Customer's Outstanding Amount.xlsx")


# groupby invoice_df by NAME and get the quantity of each item 
invoice_df = invoice_df.groupby("CUSTOMERNAME")

# get the list of customers
customers = invoice_df.groups.keys()


customer_lst = [key for key in customers]
print("Customers: ",customer_lst)
# get the last customer
last_customer = customer_lst[-1]
first_customer = customer_lst[0]    


output_COLUMNS = ["DATE","ITEMNAME","QUANTITY","PRICE","DEBIT","CREDIT"]
final_df = pd.DataFrame()
# print("Len on columns: ",len(final_df.columns))

today_date = str(pd.to_datetime("now").strftime('%d/%m/%Y')).replace("/","_")
# gernerate a random number from 1 to 1000
rand_num = str(pd.to_datetime("now").minute)+ str(pd.to_datetime("now").second)

print("Today's Date: ",rand_num)

output_path = os.path.join("OUT",today_date)
file_name = output_path+"_"+str(rand_num)+".xlsx"

final_df.to_excel(file_name,index=False, sheet_name="Sheet1")

row_index_lst = []
fin_bal_lst=[]

def data_imposer(data):
 
    # get the data from columns NAME DATE ITEMNAME QUANTITY AMOUNT PRICEPERUNIT
    new_df = pd.DataFrame()
    # chnage the date format to DD/MM/YYYY
    new_df[0] = pd.to_datetime(data["DATE"]).dt.strftime('%d/%m/%Y')
    new_df[1] = data["ITEMNAME"]
    new_df[3] = data["QUANTITY"]
    new_df[4] = data["TOTAL"]
    new_df[5] = data["TOTAL"]/data["QUANTITY"]
    new_df[6] = 0
    # remove the columns
    # new_df = new_df.drop(columns=[0,1,2,3,4,5,6])

    return new_df

# write final_df to excel



for name, group in invoice_df:
    is_last_customer = False
    is_first_customer = False
    print("Name: ",name)
    i = 0 # taken for helping loop through payments and returns
    outstanding = 0.0
    # check if it is the last item in the group
    if name == last_customer:
        # print("Last Customer")
        is_last_customer = True
    if name == first_customer:
        # print("First Customer")
        is_first_customer = True
        
        
    for cust in outstanding_df["CUSTOMERNAME"]:
        cust = str(cust).upper()
        name = name.upper()
        cust = cust.strip()
        name = name.strip()
        # print("Customer: ",cust+" Name: ",name)
        if cust in name:
            outstanding = outstanding_df["CLOSINGBALANCE"][i]
            # remove the - sign
            outstanding = str(outstanding).replace("-","")
            print("Outstanding: ",outstanding)
            break
        i+=1
    
    # bal_row = ["",name,"","BALANCE B/F :",outstanding,""]

    temp_df  = data_imposer(group)

    name = name.upper()

    total_payemnt = 0.0
    total_returns = 0.0
    payments_data={} # dictionary that takes time as key and amount as value to negate duplicate payments
    i=0
    for cust in payments_df["Customer"]:
    
        cust = str(cust).upper()
        cust = cust.strip()
        name = name.strip()

        if cust == name:

            date= payments_df["Date"][i]
            date = pd.to_datetime(date).strftime('%d/%m/%Y')
            payemnt= payments_df["Amount"][i]
            payment_time = payments_df["Date"][i]
            if payment_time in payments_data and payments_data[payment_time] == payemnt:
                continue
            else:
                sample_json ={
                    payment_time:payemnt
                }
                payments_data.update(sample_json)
                total_payemnt += payemnt
                lst = [date,"Return/Payments","","","",str(payemnt)]
                print("List: ",lst)
                temp_df.loc[len(temp_df)] = lst
                
        i+=1
    
    i=0
    for cust in returns_df["Customer Name"]:
        # cust =customer["Customer"]
        cust = str(cust).upper()
        # trim the name
        cust = cust.strip()
        name = name.strip()
        if cust == name:
            # new_returns_df = pd.DataFrame(columns=output_COLUMNS)
            date= returns_df["Date"][i]
            date = pd.to_datetime(date).strftime('%d/%m/%Y')
            returns =returns_df["Return Amount"][i]
            payment_time = returns_df["Date"][i]
            if payment_time in payments_data and payments_data[payment_time] == returns:
                continue
            else:
                sample_json ={
                    payment_time:returns
                }
                payments_data.update(sample_json)
                lst = [date,"Return/Payments","","","",returns]
                temp_df.loc[len(temp_df)] = lst
                total_returns += returns    
            
        i+=1
    # add one row for the total amount
    total_amount = float(group["TOTAL"].sum())
    
    total_amount = float(outstanding)+total_amount - (total_returns + total_payemnt)

    temp_df['date_time'] = pd.to_datetime(temp_df[0], format='%d/%m/%Y')



    temp_df = temp_df.sort_values(by="date_time", ignore_index=True, ascending=True)
    print(temp_df)
    # remove the date_time column
    temp_df = temp_df.drop(columns=['date_time'])
    name_row = ["",name,"","","",""]

    temp_df.loc[0]= name_row
    temp_df.loc[1]= output_COLUMNS
    row = ["","","","BALANCE B/F",outstanding,""]
    temp_df.loc[2] = row
    
    row =["","","","BALANCE :",total_amount, ""]

    temp_df.loc[len(temp_df)]= row


    row= ["","","","","", ""]
    temp_df.loc[len(temp_df)]= row
    row= ["","","","","", ""]
    temp_df.loc[len(temp_df)]= row

    if is_first_customer:
        row_index=3
    else:
        row_index = len(final_df)+3
    row_index_lst.append(row_index)


    if is_last_customer:
        last_table_len = len(temp_df)


    final_df = pd.concat([final_df,temp_df],ignore_index=True)

    final_df.to_excel(file_name,index=False, sheet_name="Sheet1")


colors_lst = ['00ADD8E6', '00BBFF33',
        '00E2BD00', '0000FF00', '00660066'] 
fillers = []
for color in colors_lst:
    fill = PatternFill(patternType='solid', fgColor=color)
    fillers.append(fill)



for row_index in row_index_lst:

    colour_row = ['A'+str(row_index),'B'+str(row_index),'C'+str(row_index),'D'+str(row_index),'E'+str(row_index),'F'+str(row_index)]
    column_row = ['A'+str(row_index-1),'B'+str(row_index-1),'C'+str(row_index-1),'D'+str(row_index-1),'E'+str(row_index-1),'F'+str(row_index-1)]
    bal_row = ["A"+str(row_index+1),'B'+str(row_index+1),'C'+str(row_index+1),'D'+str(row_index+1),'E'+str(row_index+1),'F'+str(row_index+1)]
    
    allign_row =['B'+str(row_index),'D'+str(row_index),'E'+str(row_index),'F'+str(row_index)]
    deb_cred_row = ['E'+str(row_index+1),'F'+str(row_index+1)]

    wb = load_workbook(file_name)
    ws= wb['Sheet1']

    for i in range(0,len(colour_row)):
        ws[colour_row[i]].fill = fillers[0]
        
    for i in range(0,len(bal_row)):
        if i==4:
            ws[bal_row[i]].fill = fillers[2]

        

    for i in range(0,len(allign_row)):
        ws[allign_row[i]].alignment = Alignment(horizontal='center')

    

    for i in range(0,len(column_row)):
    
        ws[column_row[i]].alignment = Alignment(horizontal='center')
        ws[column_row[i]].fill = fillers[0]
        # ws[column_row[i]].font = Font(b=True)
 
    for i in range(0,len(deb_cred_row)):
        ws[deb_cred_row[i]].alignment = Alignment(horizontal='center')


    wb.save(file_name)
    wb.close()
    fin_bal_lst.append(row_index-4)



# remove the first cell from the list
fin_bal_lst.pop(0)

fin_bal_lst.append(row_index_lst[-1]+last_table_len)

for row_index in fin_bal_lst:
    col_row = 'E'+str(row_index)
    wb = load_workbook(file_name)
    ws= wb['Sheet1']
    ws[col_row].fill = fillers[2]

    wb.save(file_name)
    wb.close()

for idx in range(0,len(row_index_lst)):
    first_row_index = row_index_lst[idx]-1
    last_row_index = fin_bal_lst[idx]
    wb = load_workbook(file_name)
    ws= wb['Sheet1']
    for i in range(first_row_index,last_row_index):
        for j in range(3,7):
            ws.cell(row=i, column=j).alignment = Alignment(horizontal='center')

    wb.save(file_name)
    wb.close()


# print("Count of colour rows: ",len(row_index_lst))
# print("Count of final balance rows: ",len(fin_bal_lst))
# print("NEW  Final Balance List: ",fin_bal_lst)
# print("Row Index List: ",row_index_lst)
# print("Last Table Length: ",last_table_len)


workbook = Workbook()

workbook.LoadFromFile(file_name)
worksheet = workbook.Worksheets[0]

wb = load_workbook(file_name)
ws= wb['Sheet1']
for i in range(0,len(row_index_lst)):
    first_row_index = row_index_lst[i]-1
    last_row_index = fin_bal_lst[i]
    cell_range = worksheet.Range["A"+str(first_row_index)+":F"+str(last_row_index)]
    cell_range.BorderAround(LineStyleType.Thick, Color.get_Black())
    cell_range.BorderInside(LineStyleType.Thin, Color.get_Black())
    
    workbook.SaveToFile(file_name)
    # align the text to the center


  
workbook.Dispose()

